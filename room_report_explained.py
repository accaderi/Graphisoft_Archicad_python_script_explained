# Import archicad connection (required), handle_dependencies is not necessary.
from archicad import ACConnection, handle_dependencies
# Import os for file operations, sys is unused, uuid for uuid generation.
# Note: sys is not used in this code. 
import os, sys, uuid

# Check if the importable (installed) if not returns an error
handle_dependencies('openpyxl')

# Import load_workbook from openpyxl for excel file operations.
# https://openpyxl.readthedocs.io/en/stable/index.html
# Note: Workbook class is not used in this code.
from openpyxl import Workbook, load_workbook

# Establish the connection with the Archicad software, Archicad must be open and the pln file must be open too.
#
# We can use the acu.OpenFile() utility but we need an established connection first so we need to open Archicad and a new plan
# as a minimum because all utilities use the connection.
conn = ACConnection.connect()
assert conn

# Create shorts of the commands, types and utilities.
acc = conn.commands
act = conn.types
acu = conn.utilities

# Getting the actual dirname as scriptFolder variable.
# The use of realpath is to get the canonical path adn ignore symbolic links.
scriptFolder = os.path.dirname(os.path.realpath(__file__))

# original comment -> ################################ CONFIGURATION #################################
# Define the output folder as cwd and output file in the cwd.
outputFolder = scriptFolder
outputFileName = "Room Report.xlsx"
# Define the template folder as cwd and template file in the cwd.
templateFolder = scriptFolder
templateFileName = "RDS template.xlsx"
# Create a dictionary with the cell index and initial values.
cellAddressPropertyUserIdTable = {
    "C2": act.BuiltInPropertyUserId("Zone_ZoneName"),
    "G2": act.BuiltInPropertyUserId("Zone_ZoneNumber"),
    "G3": act.BuiltInPropertyUserId("Zone_ZoneCategoryCode"),
    "G6": act.BuiltInPropertyUserId("Zone_NetArea"),
    "G7": act.BuiltInPropertyUserId("General_NetVolume"),
    "G4": act.UserDefinedPropertyUserId(["WINDOW RATE (Expression)", "Window rate calculated"]),
    "G15": act.UserDefinedPropertyUserId(["ZONES", "Temperature Requirement"]),
    "G16": act.UserDefinedPropertyUserId(["ZONES", "Illuminance Requirement"])
}
# Define the classification system.
# To get the classificationSystemName we can use the following command:
# acc.GetClassificationSystems(acc.GetClassificationSystemIds())[0].classificationSystem.name
classificationSystemName = "ARCHICAD Classification"

# Define the cells where we want to insert the different type of information. 
# C4 cell is the classification of the element
insertClassificationTo = "C4"
# C6-C16 related zones (connected zones)
insertRelatedZonesTo = ["C" + str(row) for row in range(6, 16)]
# B20-B57 equipment names
insertEquipmentNamesTo = ["B" + str(row) for row in range(20, 57)]
# D20-D57 equipment quantities
insertEquipmentQuantitiesTo = ["D" + str(row) for row in range(20, 57)]
# F20-F57 opening names
insertOpeningNamesTo = ["F" + str(row) for row in range(20, 57)]
# G20-G57 opening element ids
insertOpeningElementIDsTo = ["G" + str(row) for row in range(20, 57)]
# original comment -> ################################################################################

# This functions returns a dictionary with the elements guids : elements' details' ids, key:value pairs.
# Taking as argument: elements guid list.
def getElementsClassificationDictionary(elements):
    # Getting the guids of the elements (zones) classification
    classificationIdObjects = acc.GetClassificationsOfElements(
        elements, [acu.FindClassificationSystem(classificationSystemName)])

    # This function takes out the element's classification id from the ClassificationIdsOrErrorsWrapper object.
    def unwrapId(classification):
        if classification.classificationIds[0].classificationId.classificationItemId:
            return classification.classificationIds[0].classificationId.classificationItemId
        # If there is no id it generates one.
        else:
            return act.ClassificationItemId(uuid.uuid1())

    # List with the elements' classification ids (in our case these are the same for all three elements)
    classificationItemIds = [unwrapId(c) for c in classificationIdObjects]
    # Get the details of the classifications (guid, id, name, description).
    classificationDetails = acc.GetDetailsOfClassificationItems(classificationItemIds)

    # Getting the classification id from the classification details.
    def unwrapDetail(details):
        # If there is no attribute, meaning there is an 'error' attribute, use "<Unclassified>" id.
        if hasattr(details, "error"):
            return "<Unclassified>"
        return details.classificationItem.id

    return dict(zip(elements, [unwrapDetail(c) for c in classificationDetails]))

# This function returns the elements of the 'ElementsWrapper' created by the 'GetElementsRelatedToZones' command.
def unwrapElements(elementsWrapper):
    return elementsWrapper.elements

# This function preapres the adjacent rooms dictionary
# with the room guid:[List of the adjacent rooms guid] key:value pairs
def getAdjacentRooms(rooms):
    # Getting the adjacent elements (walls) of the zones.
    rawBoundaryObjects = acc.GetElementsRelatedToZones(rooms, ["Wall"])
    # Creating a dictionary of the rooms (guids) and their adjacent walls.
    # We are using the 'map' method to map (convert) the 'ElementsWrapper' to the 'unwrapElements' function
    # in order to get the elements (containing the guids). 
    boundaryObjects = dict(zip(rooms, list(map(unwrapElements, rawBoundaryObjects))))

    # This function is getting the guids out of the elements created above.
    def getGuid(elementIdArrayItem: act.ElementIdArrayItem) -> str:
        return str(elementIdArrayItem.elementId.guid)

    # Creating a dictionary of the rooms and a list of its adjacent walls ids as key:value pairs.
    # This step could be handled directly in the unwrapElements function
    # to return 'elementsWrapper.elements[i].elementId.guid' using list comprehension or dict comprehension.
    boundaryObjectsIds = dict({k: list(map(getGuid, v))
                                for k, v in boundaryObjects.items()})
    # Create adjacent rooms dictionary.
    adjacentRooms = {}
    # This loop is checking all the rooms if they are adjacent:
    # 1. Taking the first room.
    # 2. Preparing an empty list in the adjacentRooms list for this room adjacent rooms.
    # 3. Taking all the rooms except the same as room1 of the rooms
    # 4. With the set.instercetion method checking all the two actual room boundary walls if there is any parity
    # 5. If yes the two rooms are adjacent so append room 2 to the room1 adjacent rooms list
    # 6. Go to next room in the rooms list
    for room1 in rooms:
        adjacentRooms[room1] = []
        for room2 in rooms:
            if room1 is not room2 and set(boundaryObjectsIds[room2]).intersection(set(boundaryObjectsIds[room1])):
                adjacentRooms[room1].append(room2)

    return adjacentRooms

# This function is getting all the library parts' names in every room.
# Returns a dictionary with room ElementIdArrayItem (guid): List of the library parts' names in the room key:value pairs.
def getObjectLibPartsInRooms(rooms):
    # Get the all the elements in the room.
    elementInRooms = acc.GetElementsRelatedToZones(rooms, ["Object"])
    # Creating a dictionary with the room ElementIdArrayItem (guid): List of the library parts' guids in the room key:value pairs.
    roomElements = dict(zip(rooms, list(map(unwrapElements, elementInRooms))))
    # Getting the guid of the 'General_LibraryPartName'.
    libPartNamePropertyId = acu.GetBuiltInPropertyId('General_LibraryPartName')
    # Getting the property values ('General_LibraryPartName') of all the items in all the rooms.
    # Note: Prepare 1 list containing all the elements of a nested list
    # we can sum all the nested lists with an empty list.
    propertyValuesDictionary = acu.GetPropertyValuesDictionary(sum(roomElements.values(), []), [libPartNamePropertyId])
    return dict({room: [propertyValuesDictionary[e][libPartNamePropertyId] for e in roomElements[room]] for room in rooms})

# This function is getting all the openings' names in every room.
# Returns a dictionary with room ElementIdArrayItem (guid): [List of the openings' names, General elementID zipped as tuples]
# in the room key:value pairs.
def getOpeningsInRooms(rooms):
    # Get the all the elements in the room.
    elementInRooms = acc.GetElementsRelatedToZones(rooms, ["Door", "Window", "Skylight", "Opening"])
    # Creating a dictionary with the room ElementIdArrayItem (guid): List of the openings' guids in the room key:value pairs.
    roomElements = dict(zip(rooms, list(map(unwrapElements, elementInRooms))))
    # Getting the guid of the 'General_LibraryPartName'.
    libPartNamePropertyId = acu.GetBuiltInPropertyId('General_LibraryPartName')
    # Getting the guid of the 'General_ElementID'.
    elementIdPropertyId = acu.GetBuiltInPropertyId('General_ElementID')
    # Getting the property values ('General_LibraryPartName' and 'General_ElementID') of all the items in all the rooms.
    # Note: Prepare 1 list containing all the elements of a nested list
    # we can sum all the nested lists with an empty list.
    propertyValuesDictionary = acu.GetPropertyValuesDictionary(sum(roomElements.values(), []), [libPartNamePropertyId, elementIdPropertyId])
    return dict({room: list(zip([propertyValuesDictionary[e][libPartNamePropertyId] for e in roomElements[room]],
                                [propertyValuesDictionary[e][elementIdPropertyId] for e in roomElements[room]]))
                            for room in rooms})

# Create the WorkBookFiller class which will be handling all the excel file operations.
class WorkBookFiller:
    # Init the class with requiredd arguments: templatePath, and rooms.
    def __init__(self, templatePath, rooms):
        self.templatePath = templatePath
        self.rooms = rooms
        # Create the 'cellValuesForRoom' dictionary to write all single type details of the room. e.g. {'C4':{room:id}}.
        self.cellValuesForRoom = {}
        # Create the 'cellValueRangeForRoom' dictionary to write all list type details of the room. e.g. {['C6'-'C15']]{room:[list of adjacent rooms]}}
        self.cellValueRangeForRoom = {}
        self.zoneNumberPropertyId = acu.GetBuiltInPropertyId('Zone_ZoneNumber')
        self.zoneNamePropertyId = acu.GetBuiltInPropertyId('Zone_ZoneName')
        self.propertyValuesDictionary = acu.GetPropertyValuesDictionary(self.rooms, [self.zoneNumberPropertyId, self.zoneNamePropertyId])
        self.rooms = sorted(self.rooms, key=lambda r: self.propertyValuesDictionary[r][self.zoneNumberPropertyId])
    
    # This function does all the excel file operations using the openpyxl module.
    def SaveWorkbook(self, outputPath):
        # Initialise the  workbook.
        workbook = self._initWorkBook()
        # Fill out the cells in the workbook
        self._fillWorkbook(workbook)
        # Save the workbook to the outputPath
        workbook.save(outputPath)

    # This function prepares every celladdress data with all the rooms related property.
    # Arguments: dictionary of celladdresses as keys and propertyids as values. 
    def InsertPropertyValuesTo(self, cellAddressPropertyIdTable):
        # Create a dictionary with room: all celladdress property ids key:value pairs.
        propertyValuesDictionary = acu.GetPropertyValuesDictionary(self.rooms, list(cellAddressPropertyIdTable.values()))
        # Take each celladdress with its property id and add to each cell to the 'cellValuesForRoom' dictionary:
        # the key is the cellAddress, the value is a dictionary where the keys are the rooms and the values are their propertyvalues
        # if the propertyvalue is in the valuedictionary of the actual room.
        for cellAddress, propertyId in cellAddressPropertyIdTable.items():
            self.cellValuesForRoom[cellAddress] = {room: valuesDictionary[propertyId] for room, valuesDictionary in propertyValuesDictionary.items() if propertyId in valuesDictionary}

    # This function creates a dictionary {room:id}.
    # Inserting it to the 'self.cellValueRangeForRoom' with the proper celladdress key ['C4'].
    def InsertClassificationTo(self, cellAddress):
        self.cellValuesForRoom[cellAddress] = getElementsClassificationDictionary(self.rooms)

    # This function creates a dictionary {room: [strings contains number and value]}.
    # Inserting it to the 'self.cellValueRangeForRoom' with the proper celladdress keys list: {room: [strings contains number and value]}. 
    def InsertRelatedZonesTo(self, cellAddresses):
        # Getting the adjacent rooms dictionary (room: adjacent rooms list)
        adjacentRooms = getAdjacentRooms(self.rooms)
        adjacentRoomIds = {}
        # Fill the adjacentRoomIds dictionary with room:[list of strings conatins the adjacent rooms number and name separeted with '-'].
        for k, v in adjacentRooms.items():
            adjacentRoomIds[k] = [self.propertyValuesDictionary[item][self.zoneNumberPropertyId] + " - " + self.propertyValuesDictionary[item][self.zoneNamePropertyId] for item in v]
        
        # Add CellAddresses ['C6'-'C15'] as key and adjacentRoomIds {room:[number-name]}.
        self.cellValueRangeForRoom[tuple(cellAddresses)] = adjacentRoomIds

    # This function creates two dictionaries {room: the Library parts' names} and {room: library parts' quantities per room}.
    # Inserting these to the 'self.cellValueRangeForRoom' with the proper celladdress list as keys: {room:names}, {room:qtyties} respectively.
    def InsertObjectLibPartsTo(self, namesCellAddresses, countsCellAddresses):
        # Use the function 'getObjectLibPartsInRooms' and get a dictionary {room: [library parts' names]}.
        libpartsInRooms = getObjectLibPartsInRooms(self.rooms)
        libpartNamesInRooms = {}
        libpartCountsInRooms = {}
        # k is the room guid, v is the list with the library parts names in the room.
        for k, v in libpartsInRooms.items():
            # Making an alphabetically sorted list using set which eliminates all duplicates.
            # The list will contain the unique library part names only. 
            libpartNamesInRooms[k] = sorted(list(set(v)))
            # Count the different library parts in the room,
            # using list comprehension and 'v'as the list of all lements name and the unique libpart names list.
            libpartCountsInRooms[k] = [v.count(libpartName) for libpartName in libpartNamesInRooms[k]]

        # Add CellAddresses ['B20'-'B56'] as key and another dictionary {room:List of unique library parts' names} as value.
        self.cellValueRangeForRoom[tuple(namesCellAddresses)] = libpartNamesInRooms
        # Add CellAddresses ['D20'-'D56'] as key and another dictionary {room: List of library part quantities} as value.
        self.cellValueRangeForRoom[tuple(countsCellAddresses)] = libpartCountsInRooms

    # This function creates two dictionaries {room: the openings' names} and {room: openings' General_ElementIDs}.
    # Inserting these to the 'self.cellValueRangeForRoom' with the proper celladdress list as keys: {room:names}, {room:ids} respectively.
    def InsertOpeningsTo(self, namesCellAddresses, idsCellAddresses):
        # Use the function 'getOpeningsInRooms' and get a dictionary {room: [opening names, General_ElementIDs zipped as tuples]}.
        openingsInRooms = getOpeningsInRooms(self.rooms)
        libpartNamesInRooms = {}
        elementIdsInRooms = {}
        # k is the room guid, v is the list with the opening names and their Genral element ids.
        for k, v in openingsInRooms.items():
            # Making a list sorted by General_ElementID.
            openings = sorted(v, key=lambda t: t[1])
            # Getting the names of the openings to a list
            # and add to the 'libpartNamesInRooms' dictionary as value with the room guid as key. 
            libpartNamesInRooms[k] = [t[0] for t in openings]
            # Getting the Genral element ids of the openings to a list
            # and add to the 'elementIdsInRooms' dictionaryas value with the room guid as key.
            elementIdsInRooms[k] = [t[1] for t in openings]

        # Add CellAddresses ['F20'-'F56'] as key and another dictionary {room:List of all the openings' names} as value.
        self.cellValueRangeForRoom[tuple(namesCellAddresses)] = libpartNamesInRooms
        # Add CellAddresses ['G20'-'G56'] as key and another dictionary {room:List of all the openings' General_ElementIDs} as value.
        self.cellValueRangeForRoom[tuple(idsCellAddresses)] = elementIdsInRooms

    # This function load the template workbook for editing. Returns the loaded workbook.
    def _initWorkBook(self):
        # Using the load_workbook method to load workbook.
        # (https://openpyxl.readthedocs.io/en/stable/tutorial.html#loading-from-a-file)
        workbook = load_workbook(self.templatePath)
        # The workbook will not be a template.
        workbook.template = False
        return workbook

    def _fillWorkbook(self, workbook):
        # Get the first worksheet of the template workbook.
        base = workbook.active
        # The main loop to create all worksheets for the rooms and fill out all the cells with data.
        for room in self.rooms:
            # Getting the zone name of the actual room into the 'roomName' variable.
            roomName = self.propertyValuesDictionary[room][self.zoneNamePropertyId].replace("/", "-")
            # Getting the zone number of the actual room into the 'roomId' variable.
            roomId = self.propertyValuesDictionary[room][self.zoneNumberPropertyId].replace("/", "-")
            # If room is among the rooms (this if statement wouldn't be necessary for this code):
            if self.propertyValuesDictionary[room][self.zoneNumberPropertyId]:
                # Copy the base worksheet in the same workbook. 
                worksheet = workbook.copy_worksheet(base)
                # Set the title as a string of the actual room id and room name. e.g. '01 Bedroom' 
                worksheet.title = f"{roomId} {roomName}"
                # Looping through cellValuesForRoom dictionary.
                for cellAddress, cellValueForRoom in self.cellValuesForRoom.items():
                    if room in cellValueForRoom:
                        # Write the actual room id in the 'C4' cell.
                        worksheet[cellAddress] = cellValueForRoom[room]
                        # Print the result to the concole.
                        print(f"{worksheet.title}!{cellAddress}={cellValueForRoom[room]}")
                # Looping through cellValueRangeForRoom dictionary.
                for cellAddressrange, cellvalue in self.cellValueRangeForRoom.items():
                    # Looping through the actual zipped 'cellAddressrange' and 'cellvalue[room]'.
                    # Taking the cell address range (e.g. ['B20'-'B56'])
                    # and zipping to the cellvalues of the actual room (Names of the library parts).
                    # Note: Zip function is zipping till the last element of the shortest list.
                    # The length of the zipped list of tuples will be the same as the shortest list.
                    for cellAddress, value in zip(cellAddressrange, cellvalue[room]):
                        # Write to the actual cells the actual values.
                        worksheet[cellAddress] = value
                        # Print the result to the concole.
                        print(f"{worksheet.title}!{cellAddress}={value}")
        # Remove the template worksheet from the file.
        workbook.remove(base)

# We start from here:
# Create the templatepath variable path with filename with the os.path.join method.
templatePath = os.path.join(templateFolder, templateFileName)
# Create the main class.
# Arguments: templatePath, rooms = every 'Zone' type elements.
wbFiller = WorkBookFiller(templatePath, acc.GetElementsByType("Zone"))

# Fill the 'self.cellValuesForRoom' dictionary with the celladdress and
# the corresponding propertyvalue of the room if it exist.
# We need the PropertyIds for this function since we have 'UserIds' defined in the 'cellAddressPropertyUserIdTable'. 
wbFiller.InsertPropertyValuesTo(dict(zip(
    list(cellAddressPropertyUserIdTable.keys()),
    acc.GetPropertyIds(list(cellAddressPropertyUserIdTable.values()))
)))

# Insert related zones to the 'self.cellValueRangeForRoom' dictionary.
wbFiller.InsertRelatedZonesTo(insertRelatedZonesTo)
# Insert related library parts per room to the 'self.cellValueRangeForRoom' dictionary.
wbFiller.InsertObjectLibPartsTo(insertEquipmentNamesTo, insertEquipmentQuantitiesTo)
# Insert related openings per room to the 'self.cellValueRangeForRoom' dictionary.
wbFiller.InsertOpeningsTo(insertOpeningNamesTo, insertOpeningElementIDsTo)
# Insert related classification per room to the 'self.cellValueRangeForRoom' dictionary.
wbFiller.InsertClassificationTo(insertClassificationTo)

# Create the output path with joining the iutputfolder and output filename.
outputPath = os.path.join(outputFolder, outputFileName)
# This function not only saves the file but calling the other functions to do all excel operations.
wbFiller.SaveWorkbook(outputPath)
# Using the Archicad API 'OpenFile' utility open the excel file with the default application
# for this type of files defined in the OS.
acu.OpenFile(outputPath)

# If the file saved successfully print out to the console the ok message.
if os.path.exists(outputPath):
    print("Saved Room Report")
