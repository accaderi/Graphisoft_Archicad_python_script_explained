# Import archicad connection (required), handle_dependencies is not necessary.
from archicad import ACConnection, handle_dependencies
# Import typing module list not necessary.
from typing import List
# Import os for file operations. Sys not used.
import os, sys

# Check if the importable (installed) if not returns an error.
handle_dependencies('openpyxl')

# Import Workobook from openpyxl for excel file operations.
# https://openpyxl.readthedocs.io/en/stable/index.html
from openpyxl import Workbook

# Establish the connection with the Archicad software, Archicad must be open and the pln file must be open too.
#
# We can use the acu.OpenFile() utility but we need an established connection first so we need to open Archicad and a new plan
# as a minimum because all utilities use the connection.
conn = ACConnection.connect()
assert conn

# Create shorts of the commands, types and utilities.
acc = conn.commands
act = conn.types
acu = conn.utilities

# Getting the actual dirname as scriptFolder variable.
# The use of realpath is to get the canonical path adn ignore symbolic links.
scriptFolder = os.path.dirname(os.path.realpath(__file__))

# original comment -> ################################ CONFIGURATION #################################
# Getting into a dictionary the {element classification ID : guid}
worksheetTitlesAndElements = {
    "Beams": acc.GetElementsByType("Beam"),
    "Walls": acc.GetElementsByType("Wall")
}
# Getting the built in property user ids of the required properties into a list.
propertyUserIds = [
    act.BuiltInPropertyUserId("General_ElementID"),
    act.BuiltInPropertyUserId("General_Height"),
    act.BuiltInPropertyUserId("General_Width"),
    act.BuiltInPropertyUserId("General_Thickness")
]
outputFolder = scriptFolder
# Define the output filename.
outputFileName = "BeamAndWallGeometry.xlsx"
# original comment -> ################################################################################


# This function fit the cells to the longest value in the columns.
def AutoFitWorksheetColumns(ws):
    # Looping through the cells per column of each columns.
    for columnCells in ws.columns:
        # Getting the max string length of the filled cells using list comprehension.
        length = max(len(str(cell.value)) for cell in columnCells)
        # Setting the columns dimension to the max string length.
        # https://www.geeksforgeeks.org/python-adjusting-rows-and-columns-of-an-excel-file-using-openpyxl-module/
        # The column letter is getting the letter (string) index of the cell column instead of the number,
        # this is the required parameter of the column_dimension method.
        ws.column_dimensions[columnCells[0].column_letter].width = length


# This function prints out the worksheet content into the console.
def PrintWorksheetContent(ws):
    # Looping through the columns.
    for columnCells in ws.columns:
        # Looping through the cells of the actual column.
        for cell in columnCells:
            # Print the worksheet title, the column letter, cell row number, cell value.
            print(f"{ws.title}!{cell.column_letter}{cell.row}={cell.value}")


# This is the main function to fill out the excel worksheets.
def FillExcelWorksheetWithPropertyValuesOfElements(ws, propertyIds: List[act.PropertyIdArrayItem], elements: List[act.ElementIdArrayItem]):
    # Getting the elements with the same classification id and their guids and
    # the required properties and their guids into a dcitionary.
    propertyValuesDictionary = acu.GetPropertyValuesDictionary(elements, propertyIds)
    # Zipping into a list the propertyids with their values.
    propertyDefinitionsDictionary = dict(zip(propertyIds, acc.GetDetailsOfProperties(propertyIds)))

    # Create the base table
    # Cell in the row 2 and column 1 will have the value of 'Element Guid'
    ws.cell(row=2, column=1).value = "Element Guid"
    # Continue from row 3.
    row = 3
    # Loop through the propertyValuesDictionary and prepare, fill out the cells.
    for element, valuesDictionary in propertyValuesDictionary.items():
        # row 3 column 1 first element guid
        ws.cell(row=row, column=1).value = str(element.elementId.guid)
        # 'Go to' column 2
        column = 2
        # Loop through the valuesDictionary taking the ids, values and fill cells:
        for propertyId, propertyValue in valuesDictionary.items():
            # If the ctual row is 3 (at the beginning):
            if row == 3:
                # The cell of the row 1 and column 2 will be the actual property id. 
                ws.cell(row=1, column=column).value = str(propertyId.propertyId.guid)
                # Getting the property definition for the actual propertiId.
                propertyDefinition = propertyDefinitionsDictionary[propertyId].propertyDefinition
                # Row 2 column 3 write the 'group name / property definition name' string, e.g. 'General Parameters / Height'.
                ws.cell(row=2, column=column).value = f"{propertyDefinition.group.name} / {propertyDefinition.name}"
            # The actual row and column write the property value.
            # For the first iteration these are: row 3 column 2.
            ws.cell(row=row, column=column).value = propertyValue
            # 'Go to' next column.
            column += 1
        # 'Go to' next row.
        row += 1
    # Fit the columns widths to the max length cell values, calling the 'AutoFitWorksheetColumns' function.
    AutoFitWorksheetColumns(ws)
    # Print worksheet content into the console, calling the 'PrintWorksheetContent' function.
    PrintWorksheetContent(ws)

# Getting the property ids (guid) using the propertyuserids.
propertyIds = acc.GetPropertyIds(propertyUserIds)
# Creating a workbook.
wb = Workbook()
# Select the active worksheet.
ws = wb.active

# Variable to know the number of the actual loop number.
i = 0
# Loop through 'worksheetTitlesAndElements' dictionary to prepare the excel sheets for each item.
for title, elements in worksheetTitlesAndElements.items():
    # If we are in the first iteration we have the first sheet active and give it a title
    # of the actual element claccification Id.
    if i == 0:
        ws.title = title
    # If we are not in the first iteration we create a new sheet and give it the title.
    else:
        ws = wb.create_sheet(title)
    # Getting all required data of the actual element for the excel workbook.
    # Arguments: worksheet, property Ids (guid), elements (guid)
    FillExcelWorksheetWithPropertyValuesOfElements(ws, propertyIds, elements)
    # Go to the next iteration
    i += 1

# Prepare the created excel file's path with joining the Folder path and the filename.
excelFilePath = os.path.join(outputFolder, outputFileName)
# Save the workbook to the same folder as the script's.
wb.save(excelFilePath)
# Using the Archicad API 'OpenFile' utility open the excel file with the default application
# for this type of files defined in the OS.
acu.OpenFile(excelFilePath)

# If the file saved successfully print out to the console the ok message.
if os.path.exists(excelFilePath):
    print("Saved Excel")
