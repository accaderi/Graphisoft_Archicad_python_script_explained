# Import archicad connection (required), handle_dependencies is not necessary.
from archicad import ACConnection, handle_dependencies
# Import typing module (not necessary).
from typing import List, Dict, Any
# Import os for file operations. Sys not used. Uuid for uuid generation.
import os, sys, uuid
# Check if the importable (installed) if not returns an error.
handle_dependencies('openpyxl')

# Import Workobook from openpyxl for excel file operations.
# https://openpyxl.readthedocs.io/en/stable/index.html
from openpyxl import load_workbook

# Establish the connection with the Archicad software, Archicad must be open and the pln file must be open too.
#
# We can use the acu.OpenFile() utility but we need an established connection first so we need to open Archicad and a new plan
# as a minimum because all utilities use the connection.
conn = ACConnection.connect()
assert conn

# Create shorts of the commands, types and utilities.
acc = conn.commands
act = conn.types
acu = conn.utilities

# Getting the actual dirname as scriptFolder variable.
# The use of realpath is to get the canonical path adn ignore symbolic links.
scriptFolder = os.path.dirname(os.path.realpath(__file__))

# original comment -> ################################ CONFIGURATION #################################
inputFolder = scriptFolder
# Define the output filename.
inputFileName = "BeamAndWallGeometry.xlsx"
# original comment -> ################################################################################

# Prepare the created excel file's path with joining the Folder path and the filename.
excelFilePath = os.path.join(inputFolder, inputFileName)
# Load the excel file workbook as wb.
wb = load_workbook(excelFilePath)
# This list will be filled with the final element property value objects.
elemPropertyValues = []

for sheet in wb.worksheets:
	# Worksheet package doc.
    # https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html?highlight=max_worksheet
	maxCol = sheet.max_column
	maxRow = sheet.max_row
	# All the data from the excel workbook will be inserted into this dictionary. 
	newPropertyValues = {}
	# element id list (from the excel workbook).
	elementIds = []
	# property id list (from the excel workbook).
	propertyIds = []
	# Loop through the columns from 2 to the last (maxCol).
	for col in range (2, maxCol + 1):
		# Append the propertyId guid converted by the uuid method
        # from the string of the property id contained by the cell.
		propertyIds.append(act.PropertyId(uuid.UUID(sheet.cell(1, col).value)))
	# Loop through the rows from 3 to the last (maxRow).
	for row in range (3, maxRow + 1):
		# Get the rows in the excel workbook for the actual sheet into a dictionary.
        # e.g. 0:{0:'a8138ba2-01da-4193-aa14-974e9ca78483', 1:'SW-001', 2:6, 3:0.3,	4:0.3}
		newPropertyValues[row-3] = {col-2: sheet.cell(row, col).value for col in range (2, maxCol + 1)}
        # Append the elementId guid converted by the uuid method
        # from the string of the element id contained by the cell.
		elementIds.append(act.ElementId(uuid.UUID(sheet.cell(row, 1).value)))

    # Getting the property values of the elements based on the actual worksheet.
	propertyValuesOfElements = acc.GetPropertyValuesOfElements(elementIds, propertyIds)

    # Create a loop for the number of the element times.
	for ii in range(len(newPropertyValues)):
		# Create a loop for the number of the actual element's number of property times.
		for jj in range(len(newPropertyValues[ii])):
			# Try except block wouldn't be necessary.
			try:
                # Getting the property value directly from the 'propertyValuesOfElements' list.
				propertyValue = propertyValuesOfElements[ii].propertyValues[jj].propertyValue
                # Changing the old property value to the new one taking from the
                # 'newPropertyValues' dictionary using the numeric indexes we gave as keys.
				propertyValue.value = newPropertyValues[ii][jj]
                # Give it a 'normal' status.
				propertyValue.status = "normal"
                # Finally using the elementsId for the actual element, propertyId of the actual property
                # and the prepared property value create and append the element property value of the
                # actual element actual property to the 'elemPropertyValues' list otside of the loops.
				elemPropertyValues.append(act.ElementPropertyValue(elementIds[ii], propertyIds[jj], propertyValue))
            # If something went wrong continue.
			except:
				continue
# Set the created element property values to the corresponding elements in the Archicad project.
acc.SetPropertyValuesOfElements(elemPropertyValues)	

# original comment -> # Print the result
# Get the element ids from the elemPropertyValues list.
elementIds = [i.elementId for i in elemPropertyValues]
# Get the property ids from a set (to get the unique guids only) generated from the 'lemPropertyValues' list.
propertyIds = [act.PropertyId(guid) for guid in set(i.propertyId.guid for i in elemPropertyValues)]
# Creae a property values dictionary from the elementids and propertyids.
propertyValuesDictionary = acu.GetPropertyValuesDictionary(elementIds, propertyIds)
# Loop through and taking each item of the 'propertyValuesDictionary'.
for elementId, valuesDictionary in propertyValuesDictionary.items():
	# For each element of the 'propertyValuesDictionary', loop through the values dictionary
    # and take each property ids and values and print them onto the console.
    for propertyId, value in valuesDictionary.items():
        print(f"{elementId.guid} {propertyId.guid} {value}")
